"""
AI处理Excel - 后端API服务 (使用Excel MCP Server)
基于 excel-mcp-server 的Excel处理功能
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import os
import sys
import uuid
import io
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border
from datetime import datetime

app = Flask(__name__)
CORS(app)

# 配置
UPLOAD_FOLDER = 'uploads'
DOWNLOAD_FOLDER = 'downloads'
EXCEL_FILES_PATH = 'excel_files'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(DOWNLOAD_FOLDER, exist_ok=True)
os.makedirs(EXCEL_FILES_PATH, exist_ok=True)

# 添加excel_mcp_server到路径
MCP_SERVER_PATH = 'D:/excel-mcp-server-main/excel-mcp-server-main/src'
if MCP_SERVER_PATH not in sys.path:
    sys.path.insert(0, MCP_SERVER_PATH)

# 导入excel_mcp模块
try:
    from excel_mcp.workbook import create_workbook, create_sheet, get_workbook_info
    from excel_mcp.data import read_excel_range_with_metadata, write_data
    from excel_mcp.formatting import format_range
    from excel_mcp.sheet import (
        copy_sheet, delete_sheet, rename_sheet,
        insert_row, insert_cols, delete_rows, delete_cols
    )
    from excel_mcp.calculations import apply_formula
    from excel_mcp.tables import create_excel_table
    from excel_mcp.chart import create_chart_in_sheet
    from excel_mcp.pivot import create_pivot_table
    MCP_AVAILABLE = True
except ImportError as e:
    print(f"警告: 无法导入excel_mcp模块: {e}")
    MCP_AVAILABLE = False


def get_excel_operations():
    """获取可用的Excel操作列表"""
    operations = [
        {"id": "read", "name": "读取数据", "description": "读取Excel文件中的数据"},
        {"id": "write", "name": "写入数据", "description": "向Excel文件写入数据"},
        {"id": "create_workbook", "name": "创建工作簿", "description": "创建一个新的Excel工作簿"},
        {"id": "create_sheet", "name": "创建工作表", "description": "在工作簿中创建新工作表"},
        {"id": "format_cells", "name": "格式化单元格", "description": "设置单元格的格式（字体、颜色、边框等）"},
        {"id": "apply_formula", "name": "应用公式", "description": "在单元格中应用Excel公式"},
        {"id": "create_table", "name": "创建表格", "description": "创建Excel表格"},
        {"id": "create_chart", "name": "创建图表", "description": "创建数据图表"},
        {"id": "create_pivot", "name": "创建数据透视表", "description": "创建数据透视表"},
        {"id": "delete_rows", "name": "删除行", "description": "删除指定的行"},
        {"id": "delete_cols", "name": "删除列", "description": "删除指定的列"},
        {"id": "insert_rows", "name": "插入行", "description": "插入新的行"},
        {"id": "insert_cols", "name": "插入列", "description": "插入新的列"},
        {"id": "copy_sheet", "name": "复制工作表", "description": "复制工作表"},
        {"id": "delete_sheet", "name": "删除工作表", "description": "删除工作表"},
        {"id": "rename_sheet", "name": "重命名工作表", "description": "重命名工作表"},
    ]
    return operations


def process_excel_with_ai(file_path, instruction):
    """
    使用AI理解用户需求并处理Excel文件
    这里使用简单的规则匹配，实际可以接入大模型
    """
    # 读取Excel文件
    df = pd.read_excel(file_path) if file_path.endswith('.xlsx') else pd.read_csv(file_path)

    result = {
        'message': '',
        'df': df,
        'operation': None
    }

    instruction_lower = instruction.lower()

    # 规则匹配处理
    if '总计' in instruction or '总和' in instruction or 'sum' in instruction_lower:
        # 计算总和
        for col in df.columns:
            if pd.api.types.is_numeric_dtype(df[col]):
                total = df[col].sum()
                result['message'] = f'{col}列的总和是: {total}'
                result['operation'] = 'sum'
                break

    elif '平均' in instruction or 'avg' in instruction_lower:
        # 计算平均值
        for col in df.columns:
            if pd.api.types.is_numeric_dtype(df[col]):
                avg = df[col].mean()
                result['message'] = f'{col}列的平均值是: {avg}'
                result['operation'] = 'avg'
                break

    elif '最大' in instruction or 'max' in instruction_lower:
        # 计算最大值
        for col in df.columns:
            if pd.api.types.is_numeric_dtype(df[col]):
                max_val = df[col].max()
                result['message'] = f'{col}列的最大值是: {max_val}'
                result['operation'] = 'max'
                break

    elif '最小' in instruction or 'min' in instruction_lower:
        # 计算最小值
        for col in df.columns:
            if pd.api.types.is_numeric_dtype(df[col]):
                min_val = df[col].min()
                result['message'] = f'{col}列的最小值是: {min_val}'
                result['operation'] = 'min'
                break

    elif '筛选' in instruction or 'filter' in instruction_lower:
        # 筛选数据
        if '>' in instruction:
            # 提取数字
            import re
            numbers = re.findall(r'[\d.]+', instruction)
            if numbers:
                threshold = float(numbers[0])
                for col in df.columns:
                    if pd.api.types.is_numeric_dtype(df[col]):
                        df_filtered = df[df[col] > threshold]
                        result['df'] = df_filtered
                        result['message'] = f'筛选{col}大于{threshold}的行，共{len(df_filtered)}行'
                        result['operation'] = 'filter'
                        break

    elif '删除' in instruction and '重复' in instruction:
        # 删除重复行
        df_dropped = df.drop_duplicates()
        result['df'] = df_dropped
        result['message'] = f'删除了{len(df) - len(df_dropped)}行重复数据'
        result['operation'] = 'drop_duplicates'

    elif '填充' in instruction and ('空' in instruction or '0' in instruction):
        # 填充空值
        df_filled = df.fillna(0)
        result['df'] = df_filled
        result['message'] = '已将所有空值填充为0'
        result['operation'] = 'fill_na'

    elif '排序' in instruction or 'sort' in instruction_lower:
        # 排序 - 支持多种表达方式
        # 判断排序方向
        descending_keywords = ['从高到低', '降序', 'desc', '最大在前', '大到小']
        ascending_keywords = ['从低到高', '升序', 'asc', '最小在前', '小到大']

        is_descending = any(kw in instruction for kw in descending_keywords)
        is_ascending = any(kw in instruction for kw in ascending_keywords)

        # 如果没有明确指定方向，默认为降序（从高到低）
        ascending = not is_descending if is_descending or is_ascending else False

        # 找到要排序的列
        sort_column = None
        for col in df.columns:
            if pd.api.types.is_numeric_dtype(df[col]):
                sort_column = col
                break

        if sort_column:
            df_sorted = df.sort_values(by=sort_column, ascending=ascending)
            result['df'] = df_sorted
            order_text = '从低到高' if ascending else '从高到低'
            result['message'] = f'已按{sort_column}列{order_text}排序'
            result['operation'] = 'sort'
        else:
            result['message'] = '没有找到可排序的数值列'

    elif '行数' in instruction or 'count' in instruction_lower:
        # 统计行数
        result['message'] = f'共有{len(df)}行数据'
        result['operation'] = 'count'

    else:
        # 默认返回数据预览
        result['message'] = f'已读取文件，共{len(df)}行{len(df.columns)}列数据'

    return result


@app.route('/api/operations', methods=['GET'])
def get_operations():
    """获取可用的Excel操作列表"""
    return jsonify({
        'success': True,
        'operations': get_excel_operations(),
        'mcp_available': MCP_AVAILABLE
    })


@app.route('/api/process-excel', methods=['POST'])
def process_excel():
    """
    处理Excel文件的主接口
    """
    try:
        # 检查文件
        if 'file' not in request.files:
            return jsonify({'error': '未上传文件'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '文件名为空'}), 400

        # 获取用户需求
        instruction = request.form.get('instruction', '')
        if not instruction:
            return jsonify({'error': '请输入处理需求'}), 400

        # 保存上传的文件
        file_id = str(uuid.uuid4())
        file_ext = os.path.splitext(file.filename)[1]
        upload_path = os.path.join(UPLOAD_FOLDER, f"{file_id}{file_ext}")
        file.save(upload_path)

        # 处理Excel文件
        result = process_excel_with_ai(upload_path, instruction)
        df_processed = result['df']

        # 保存处理后的文件
        download_filename = f"processed_{file_id}{file_ext}"
        download_path = os.path.join(DOWNLOAD_FOLDER, download_filename)

        if file_ext == '.csv':
            df_processed.to_csv(download_path, index=False, encoding='utf-8-sig')
        else:
            df_processed.to_excel(download_path, index=False, engine='openpyxl')

        # 生成预览数据
        preview_data = df_processed.head(10).to_string()

        return jsonify({
            'success': True,
            'message': result['message'],
            'downloadUrl': f'http://localhost:5000/api/download/{download_filename}',
            'preview': preview_data,
            'operation': result['operation'],
            'stats': {
                'originalRows': len(pd.read_excel(upload_path) if file_ext != '.csv' else pd.read_csv(upload_path)),
                'processedRows': len(df_processed),
                'columns': df_processed.columns.tolist()
            }
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/read-excel-preview', methods=['POST'])
def read_excel_preview():
    """读取上传的Excel文件预览"""
    try:
        # 检查文件
        if 'file' not in request.files:
            return jsonify({'error': '未上传文件'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '文件名为空'}), 400

        # 保存临时文件
        file_id = str(uuid.uuid4())
        file_ext = os.path.splitext(file.filename)[1]
        temp_path = os.path.join(UPLOAD_FOLDER, f"temp_{file_id}{file_ext}")
        file.save(temp_path)

        # 读取Excel文件
        if file_ext == '.csv':
            df = pd.read_csv(temp_path, encoding='utf-8-sig')
        else:
            df = pd.read_excel(temp_path)

        # 删除临时文件
        try:
            os.remove(temp_path)
        except:
            pass

        return jsonify({
            'success': True,
            'data': df.to_dict(orient='records'),
            'columns': df.columns.tolist(),
            'rows': len(df)
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/save-edited-data', methods=['POST'])
def save_edited_data():
    """保存编辑后的数据为Excel文件"""
    try:
        data = request.json
        edited_data = data.get('data', [])
        filename = data.get('filename', 'edited.xlsx')

        # 转换为DataFrame
        df = pd.DataFrame(edited_data)

        # 确保文件扩展名正确
        if not filename.endswith(('.xlsx', '.xls', '.csv')):
            filename = filename.rsplit('.', 1)[0] + '.xlsx' if '.' in filename else filename + '.xlsx'

        # 生成文件ID
        file_id = str(uuid.uuid4())
        download_filename = f"edited_{file_id}_{filename}"
        download_path = os.path.join(DOWNLOAD_FOLDER, download_filename)

        # 保存文件
        if download_filename.endswith('.csv'):
            df.to_csv(download_path, index=False, encoding='utf-8-sig')
        else:
            df.to_excel(download_path, index=False, engine='openpyxl')

        return jsonify({
            'success': True,
            'message': '数据已保存',
            'downloadUrl': f'http://localhost:5000/api/download/{download_filename}'
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/read-excel', methods=['POST'])
def read_excel():
    """读取Excel文件数据"""
    try:
        data = request.json
        filepath = data.get('filepath')
        sheet_name = data.get('sheet_name')
        start_cell = data.get('start_cell', 'A1')
        end_cell = data.get('end_cell')

        if not filepath or not sheet_name:
            return jsonify({'error': '缺少必要参数'}), 400

        # 构建完整路径
        if not os.path.isabs(filepath):
            filepath = os.path.join(EXCEL_FILES_PATH, filepath)

        if not os.path.exists(filepath):
            return jsonify({'error': f'文件不存在: {filepath}'}), 404

        # 使用pandas读取
        df = pd.read_excel(filepath, sheet_name=sheet_name)

        return jsonify({
            'success': True,
            'data': df.to_dict(orient='records'),
            'rows': len(df),
            'columns': df.columns.tolist()
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/write-excel', methods=['POST'])
def write_excel():
    """向Excel文件写入数据"""
    try:
        data = request.json
        filepath = data.get('filepath')
        sheet_name = data.get('sheet_name')
        cell_data = data.get('data', [])  # 二维数组
        start_cell = data.get('start_cell', 'A1')

        if not filepath or not sheet_name:
            return jsonify({'error': '缺少必要参数'}), 400

        # 构建完整路径
        if not os.path.isabs(filepath):
            filepath = os.path.join(EXCEL_FILES_PATH, filepath)

        # 确保目录存在
        os.makedirs(os.path.dirname(filepath), exist_ok=True)

        # 使用pandas写入
        df = pd.DataFrame(cell_data)

        if os.path.exists(filepath):
            # 文件存在，追加到指定工作表
            with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                df.to_excel(writer, sheet_name=sheet_name, startrow=int(start_cell[1:])-1 if start_cell[1:].isdigit() else 0, index=False, header=False)
        else:
            # 创建新文件
            df.to_excel(filepath, sheet_name=sheet_name, index=False, engine='openpyxl')

        return jsonify({
            'success': True,
            'message': f'数据已写入 {filepath}'
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/create-workbook', methods=['POST'])
def create_workbook_api():
    """创建新的Excel工作簿"""
    try:
        data = request.json
        filepath = data.get('filepath')

        if not filepath:
            return jsonify({'error': '缺少文件名'}), 400

        # 构建完整路径
        if not os.path.isabs(filepath):
            filepath = os.path.join(EXCEL_FILES_PATH, filepath)

        # 确保目录存在
        os.makedirs(os.path.dirname(filepath), exist_ok=True)

        # 创建新工作簿
        from openpyxl import Workbook
        wb = Workbook()
        wb.save(filepath)

        return jsonify({
            'success': True,
            'message': f'已创建工作簿: {filepath}'
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/apply-formula', methods=['POST'])
def apply_formula_api():
    """在单元格中应用公式"""
    try:
        data = request.json
        filepath = data.get('filepath')
        sheet_name = data.get('sheet_name')
        cell = data.get('cell')
        formula = data.get('formula')

        if not all([filepath, sheet_name, cell, formula]):
            return jsonify({'error': '缺少必要参数'}), 400

        # 构建完整路径
        if not os.path.isabs(filepath):
            filepath = os.path.join(EXCEL_FILES_PATH, filepath)

        if not os.path.exists(filepath):
            return jsonify({'error': f'文件不存在: {filepath}'}), 404

        # 使用openpyxl应用公式
        wb = load_workbook(filepath)
        ws = wb[sheet_name]
        ws[cell] = formula
        wb.save(filepath)

        return jsonify({
            'success': True,
            'message': f'已在 {cell} 单元格应用公式: {formula}'
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/download/<filename>', methods=['GET'])
def download_file(filename):
    """下载处理后的文件"""
    try:
        file_path = os.path.join(DOWNLOAD_FOLDER, filename)
        if os.path.exists(file_path):
            return send_file(
                file_path,
                as_attachment=True,
                download_name=filename
            )
        else:
            return jsonify({'error': '文件不存在'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/health', methods=['GET'])
def health_check():
    """健康检查"""
    return jsonify({
        'status': 'ok',
        'service': 'AI Excel Processor (MCP)',
        'version': '2.0',
        'mcp_available': MCP_AVAILABLE,
        'mcp_path': MCP_SERVER_PATH
    })


if __name__ == '__main__':
    print("=" * 50)
    print("AI处理Excel - 后端服务 (基于Excel MCP Server)")
    print("=" * 50)
    print(f"Excel MCP Server: {'已安装' if MCP_AVAILABLE else '未安装'}")
    print("服务启动在: http://localhost:5000")
    print("=" * 50)
    app.run(host='0.0.0.0', port=5000, debug=True)
